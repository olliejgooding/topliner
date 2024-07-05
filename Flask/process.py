import pandas as pd
from pathlib import Path
import pyreadstat
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# List of variables to ignore
ignore_vars = [
    'Respondent_Serial_SourceFile', 'Respondent_Origin1', 'Respondent_Origin2', 'Respondent_Origin3',
    'Respondent_Origin4', 'Respondent_Origin5', 'Respondent_Origin6', 'Respondent_Origin_Other', 'Respondent_ID',
    'DataCollection_Status1', 'DataCollection_Status2', 'DataCollection_Status3', 'DataCollection_Status4',
    'DataCollection_Status5', 'DataCollection_Status6', 'DataCollection_Status7', 'DataCollection_Status8',
    'DataCollection_Status9', 'DataCollection_InterviewerID', 'DataCollection_StartTime', 'DataCollection_FinishTime',
    'DataCollection_MetadataVersionNumber', 'DataCollection_MetadataVersionGUID', 'DataCollection_RoutingContext',
    'DataCollection_Variant', 'DataCollection_EndQuestion', 'DataCollection_TerminateSignal', 'DataCollection_SeedValue',
    'DataCollection_InterviewEngine', 'DataCollection_CurrentPage', 'DataCollection_Debug', 'DataCollection_ServerTimeZone',
    'DataCollection_InterviewerTimeZone', 'DataCollection_RespondentTimeZone', 'DataCollection_BatchID',
    'DataCollection_BatchName', 'DataCollection_DataEntryMode', 'DataCollection_Removed', 'DataCollection_InterviewMode',
    'DataCleaning_Note', 'DataCleaning_Status1', 'DataCleaning_Status2', 'DataCleaning_ReviewStatus1',
    'DataCleaning_ReviewStatus2', 'DataCleaning_ReviewStatus3', 'DataCleaning_ReviewStatus4', 'projnum', 'ovquot',
    'failcodes', 'Route', 'SectionTiming_START', 'SectionTiming_SCREENER', 'SectionTiming_A', 'SectionTiming_B',
    'SectionTiming_C', 'SectionTiming_D', 'SectionTiming_E', 'SectionTiming_F', 'SectionTiming_G', 'SectionTiming_H',
    'SectionTiming_I', 'SectionTiming_J', 'SectionTiming_K', 'SectionTiming_L', 'SectionTiming_M', 'SectionTiming_N',
    'SectionTimingVolatile_START', 'SectionTimingVolatile_SCREENER', 'SectionTimingVolatile_A', 'SectionTimingVolatile_B',
    'SectionTimingVolatile_C', 'SectionTimingVolatile_D', 'SectionTimingVolatile_E', 'SectionTimingVolatile_F',
    'SectionTimingVolatile_G', 'SectionTimingVolatile_H', 'SectionTimingVolatile_I', 'SectionTimingVolatile_J',
    'SectionTimingVolatile_K', 'SectionTimingVolatile_L', 'SectionTimingVolatile_M', 'SectionTimingVolatile_N',
    'ContactPortal_con', 'ContactPortal_tel', 'WFH_contact', 'WFH_email', 'WFH_refcontact', 'WFH_refcontact_Codes',
    'maxe', 'email', 'emailContact', 'econ', 'econ_Codes', 'semail1', 'semail2', 'etemplate', 'firstline', 'subjectline',
    'BrowserCapabilities_START', 'BrowserCapabilities_END', 'gorcode', 'Qovquot'
]


# Define ignore_vars
#ignore_vars = ['var1', 'var2']  # Example ignore variables, replace with actual ones as needed

def spss_to_excel1(file_path, output_file):
    df, meta = pyreadstat.read_sav(file_path)

    # Dictionary of variable labels
    variable_labels = dict(zip(meta.column_names, meta.column_labels))
    
    # Dictionary of value labels
    value_labels = meta.value_labels
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        worksheet = writer.book.create_sheet('Sheet1')
        writer.sheets['Sheet1'] = worksheet
         # Write the header row
        headers = ['variable_name', 'variable_label', 'total_base']
        worksheet.append(headers)       
        
        current_row = 2
        first_var_total = None
        
        for var in df.columns:
            if var not in ignore_vars:
                var_label = variable_labels.get(var, var)
                var_data = df[var].dropna()
                total_count = var_data.count()
                value_counts = var_data.value_counts()
                
                # Write variable label and total count
                worksheet.cell(row=current_row, column=1, value=var)
                worksheet.cell(row=current_row, column=2, value=var_label)
                worksheet.cell(row=current_row, column=3, value=total_count)
                # Capture the total count of the first variable
                if first_var_total is None:
                    first_var_total = total_count
                
                # Retrieve value labels for the current variable
                label_key = f"labels{list(meta.column_names).index(var)-93}"
                var_value_labels = value_labels.get(label_key, {})
                
                # Write value labels and their counts
                col_idx = 4
                for value, label in var_value_labels.items():
                    worksheet.cell(row=current_row, column=col_idx, value=label)
                    #worksheet.cell(row=current_row + 1, column=col_idx, value=value_counts.get(value, 0))
                    col_idx += 1
                
                # Increment the current row for the next variable
                current_row += 1
        # Adjust column widths to the average width needed
        for col in worksheet.columns:
            total_length = 0
            cell_count = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if cell.value:
                        total_length += len(str(cell.value))
                        cell_count += 1
                except:
                    pass
            if cell_count > 0:
                average_length = total_length / cell_count
                adjusted_width = average_length + 2
                worksheet.column_dimensions[column].width = adjusted_width
        # Apply conditional formatting to highlight cells in "total_base" column
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        if first_var_total is not None:
            rule = CellIsRule(operator='equal', formula=[str(first_var_total)], fill=fill)
            worksheet.conditional_formatting.add(f'C2:C{current_row-1}', rule)


# Example usage
#spss_to_excel('your_spss_file.sav', 'output.xlsx')

# Example usage
file_path = Path(r"C:\Users\Owner\Downloads\DummyautoSPSS.sav")
output_file = Path(r'output.xlsx')

def spss_to_excelcheck(file_path, output_file):
    df, meta = pyreadstat.read_sav(file_path)

    variable_labels = dict(zip(meta.column_names, meta.column_labels))
    value_labels = meta.value_labels

    return print(value_labels,"!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!\n")


spss_to_excel1(file_path, output_file)
